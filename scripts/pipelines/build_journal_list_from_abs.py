#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "openpyxl>=3.1",
# ]
# ///
"""Build a journals.json scope file from the ABS Academic Journal Guide xlsx.

The Chartered ABS publishes the Academic Journal Guide (AJG) as an
Excel workbook each year. Most management-discipline systematic
reviews scope by ABS rank (3 / 4 / 4*) and field code (ENT-SBM /
INNOV / INFO MAN / OR&MANSCI / …). This script does the filtering once
and writes the result to a JSON file the user can paste into their
project's `search_config.py` JOURNALS dict, or import directly.

The script is one of a planned family — `build_journal_list_from_<source>.py`
— so the abstraction never silently locks the SLR to ABS as the only
journal-ranking source. Future siblings: `_from_jcr.py`, `_from_fnege.py`,
`_from_abdc.py`. Each reads its source's native format and emits the
same `journals.json` schema; downstream consumers don't care which
source was used.

ABS spreadsheet shape this script expects:

  Column    Header (case-sensitive)        Example
  -------   ----------------------------   ----------------------
  Field     Field code (ABS taxonomy)      ENT-SBM
  Journal   Full journal title             Journal of Business Venturing
  ISSN      ISSN (with or without hyphen)  0883-9026 or 08839026
  AJG 2024  Rank (free-form string)        4*  /  4  /  3  /  2

Different ABS years ship slightly different column names ("AJG 2024" vs
"AJG 2021" vs "ABS Rating"). Use --rank-column to override if your
copy uses a non-default name.

Usage:
    uv run build_journal_list_from_abs.py --abs-xlsx ABSRanking2024_Fulllist.xlsx \\
        --ranks 3,4,4* --fields ENT-SBM,INNOV --out journals.json

    # Broader scope: include OR&MANSCI top-tier and INFO MAN top-tier:
    uv run build_journal_list_from_abs.py --abs-xlsx ABSRanking2024_Fulllist.xlsx \\
        --ranks 4,4* --fields OR&MANSCI,INFO\\ MAN --out journals.json

The output `journals.json` carries:
  - rating_source / ranks_included / fields_included as audit trail
  - journals: list of {issn, title, rating, field} dicts
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

# Default header names. Override via --field-column / --title-column /
# --issn-column / --rank-column when the user's spreadsheet uses
# different labels (e.g. older AJG editions or a custom export).
DEFAULT_FIELD_COL = "Field"
DEFAULT_TITLE_COL = "Journal"
DEFAULT_ISSN_COL = "ISSN"
DEFAULT_RANK_COL = "AJG 2024"


def _normalise_issn(issn: object) -> str:
    """Strip non-digits / X, then re-insert the hyphen at position 4.

    Matches the canonicalization in `import_to_zotero._canonicalize_issn`
    so the build-time output matches the ingest-time form the search
    pipeline produces. Returns "" if the input doesn't normalize to
    8 digits + optional check-X.
    """
    if issn is None:
        return ""
    raw = str(issn).strip()
    cleaned = "".join(c for c in raw if c.isdigit() or c.upper() == "X")
    if len(cleaned) != 8:
        return ""
    return f"{cleaned[:4]}-{cleaned[4:].upper()}"


def _open_workbook(path: Path):
    """Open the xlsx with openpyxl. Raises a friendly error if missing."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            "ERROR: openpyxl not installed. Run via `uv run` so the PEP 723 "
            "block at the top of this script picks it up automatically."
        )
    if not path.is_file():
        sys.exit(f"ERROR: ABS xlsx not found: {path}")
    return load_workbook(path, data_only=True, read_only=True)


def _resolve_columns(
    header_row: tuple,
    field_col: str,
    title_col: str,
    issn_col: str,
    rank_col: str,
) -> dict[str, int]:
    """Map header strings to column indices. Errors clearly if any
    expected column is missing — keeps misconfiguration surface-able."""
    headers = [str(c.value).strip() if c.value else "" for c in header_row]
    needed = {
        "field": field_col,
        "title": title_col,
        "issn": issn_col,
        "rank": rank_col,
    }
    out: dict[str, int] = {}
    for slot, name in needed.items():
        if name not in headers:
            sys.exit(
                f"ERROR: column {name!r} not found in spreadsheet header. "
                f"Available columns: {headers}\n"
                f"Use --{slot}-column to override the expected header name."
            )
        out[slot] = headers.index(name)
    return out


def _parse_csv_arg(value: str) -> list[str]:
    """`A,B,C` → `['A', 'B', 'C']`. Whitespace-trimmed; empty entries dropped."""
    return [v.strip() for v in (value or "").split(",") if v.strip()]


def filter_journals(
    rows: list[tuple],
    columns: dict[str, int],
    *,
    ranks: list[str] | None = None,
    fields: list[str] | None = None,
) -> list[dict[str, str]]:
    """Project + filter spreadsheet rows into the journals.json shape.

    Pure function: takes already-loaded row tuples (one per spreadsheet
    row) plus the column mapping. Used by both the CLI entry point and
    the unit tests so the filter logic is exercised without a real
    xlsx file.
    """
    rank_set = {r.lower() for r in (ranks or [])}
    field_set = {f.lower() for f in (fields or [])}
    out: list[dict[str, str]] = []
    for row in rows:
        if not row or len(row) <= max(columns.values()):
            continue
        rank = str(row[columns["rank"]] or "").strip()
        field = str(row[columns["field"]] or "").strip()
        title = str(row[columns["title"]] or "").strip()
        issn = _normalise_issn(row[columns["issn"]])
        if not (title and issn):
            continue
        if rank_set and rank.lower() not in rank_set:
            continue
        if field_set and field.lower() not in field_set:
            continue
        out.append({
            "issn": issn,
            "title": title,
            "rating": f"ABS {rank}",
            "field": field,
        })
    return out


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Filter the ABS Academic Journal Guide xlsx by rank + field "
            "and emit a journals.json scope file for the SLR search pipeline."
        ),
    )
    parser.add_argument(
        "--abs-xlsx", required=True,
        help="Path to the ABS AJG xlsx (e.g. ABSRanking2024_Fulllist.xlsx).",
    )
    parser.add_argument(
        "--ranks", default="",
        help=(
            "Comma-separated list of rank strings to include (e.g. "
            "`3,4,4*`). Match is case-insensitive against the rank "
            "column. Empty = no rank filter."
        ),
    )
    parser.add_argument(
        "--fields", default="",
        help=(
            "Comma-separated list of ABS field codes to include (e.g. "
            "`ENT-SBM,INNOV`). Empty = no field filter. Watch shell "
            "escaping: `OR&MANSCI` needs quoting in bash."
        ),
    )
    parser.add_argument(
        "--out", default="journals.json",
        help="Output path (default: journals.json in the current directory).",
    )
    parser.add_argument(
        "--field-column", default=DEFAULT_FIELD_COL,
        help=f"Header for the field-code column (default: {DEFAULT_FIELD_COL!r}).",
    )
    parser.add_argument(
        "--title-column", default=DEFAULT_TITLE_COL,
        help=f"Header for the journal-title column (default: {DEFAULT_TITLE_COL!r}).",
    )
    parser.add_argument(
        "--issn-column", default=DEFAULT_ISSN_COL,
        help=f"Header for the ISSN column (default: {DEFAULT_ISSN_COL!r}).",
    )
    parser.add_argument(
        "--rank-column", default=DEFAULT_RANK_COL,
        help=(
            f"Header for the rank column (default: {DEFAULT_RANK_COL!r}). "
            "Older AJG editions use 'AJG 2021' / 'AJG 2018'; some custom "
            "exports use 'ABS Rating'."
        ),
    )
    parser.add_argument(
        "--sheet", default="",
        help="Sheet name (default: first sheet in the workbook).",
    )
    args = parser.parse_args()

    abs_path = Path(args.abs_xlsx)
    out_path = Path(args.out)
    ranks = _parse_csv_arg(args.ranks)
    fields = _parse_csv_arg(args.fields)

    wb = _open_workbook(abs_path)
    sheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows()
    try:
        header_row = next(rows_iter)
    except StopIteration:
        sys.exit(f"ERROR: spreadsheet {abs_path} is empty.")
    columns = _resolve_columns(
        header_row,
        args.field_column, args.title_column,
        args.issn_column, args.rank_column,
    )
    # Materialize row tuples (read-only mode iter_rows yields cells).
    data_rows = [tuple(c.value for c in row) for row in rows_iter]

    journals = filter_journals(
        data_rows, columns, ranks=ranks, fields=fields,
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps({
        "rating_source": "ABS",
        "rank_column": args.rank_column,
        "ranks_included": ranks,
        "fields_included": fields,
        "journal_count": len(journals),
        "journals": journals,
    }, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(
        f"Wrote {len(journals)} journals to {out_path} "
        f"(ranks={ranks or 'all'}, fields={fields or 'all'}).",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
